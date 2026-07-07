# Builds 40_PHASE4_PORTFOLIO/BENCHMARK_WORKBOOK.xlsx (values only) per MISSION_BRIEF P4
# and assumption A16. Run from the mission root: python 05_STATE/tools/build_workbook.py
import csv, json, os, sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = os.getcwd()
ASOF = "2026-07-07"
DISC = ("Educational benchmark for a paper-trading learning exercise. "
        "Simulated allocations, not investment advice or recommendations.")

def read_csv_skip_disclaimer(path):
    with open(path, encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if rows and rows[0] and rows[0][0].startswith("Educational benchmark"):
        rows = rows[1:]
    return rows

def add_sheet(wb, title, note, rows):
    ws = wb.create_sheet(title=title)
    ws.append([note])
    ws["A1"].font = Font(italic=True, size=9)
    for r in rows:
        out = []
        for cell in r:
            if isinstance(cell, str):
                s = cell.strip()
                try:
                    out.append(float(s) if ("." in s or "e" in s.lower()) else int(s))
                    continue
                except ValueError:
                    pass
                out.append(cell)
            else:
                out.append(cell)
        ws.append(out)
    hdr_row = 2
    for c in range(1, ws.max_column + 1):
        ws.cell(row=hdr_row, column=c).font = Font(bold=True)
    widths = {}
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 60)):
        for cell in row:
            if cell.value is not None:
                w = min(len(str(cell.value)), 60)
                widths[cell.column] = max(widths.get(cell.column, 8), w + 2)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"
    return ws

wb = Workbook()
wb.remove(wb.active)

note = f"{DISC} | Data as-of {ASOF} (quotes: 2026-07-07 A-share close; filings: FY2025/Q1-2026)."

uni = read_csv_skip_disclaimer("10_PHASE1_VERIFICATION/VERIFIED_WATCHLIST.csv")
add_sheet(wb, "Verified_Universe", note, uni)

fun = read_csv_skip_disclaimer("20_PHASE2_FUNDAMENTALS/FUNDAMENTALS.csv")
add_sheet(wb, "Fundamentals", note + " | Units: CNY millions; percentages plain.", fun)

val = read_csv_skip_disclaimer("30_PHASE3_VALUATION_SCORING/VALUATION.csv")
add_sheet(wb, "Valuation", note, val)

sco = read_csv_skip_disclaimer("30_PHASE3_VALUATION_SCORING/SCORES.csv")
add_sheet(wb, "Scores", note + " | FINAL post-red-team scores (see REDTEAM_RESPONSES.md).", sco)

por = read_csv_skip_disclaimer("40_PHASE4_PORTFOLIO/BENCHMARK_PORTFOLIO.csv")
add_sheet(wb, "Portfolio", note + " | Fictional 100,000 CNY simulation.", por)

with open("90_BIBLIOGRAPHY/sources.json", encoding="utf-8") as f:
    src = json.load(f)
keys = ["id", "url", "title", "publisher", "date", "tier", "lang", "accessed", "verified"]
rows = [keys + ["used_in"]]
for e in src:
    rows.append([e.get(k, "") for k in keys] + [";".join(e.get("used_in", []))])
add_sheet(wb, "Sources", note + f" | {len(src)} ledger entries.", rows)

out = "40_PHASE4_PORTFOLIO/BENCHMARK_WORKBOOK.xlsx"
wb.save(out)

wb2 = load_workbook(out)  # round-trip test per Gate G4
sheets = wb2.sheetnames
dims = {s: (wb2[s].max_row, wb2[s].max_column) for s in sheets}
expected = ["Verified_Universe", "Fundamentals", "Valuation", "Scores", "Portfolio", "Sources"]
ok = sheets == expected
print("sheets:", sheets)
print("dims:", dims)
print("ROUNDTRIP:", "PASS" if ok else "FAIL (tab order/name mismatch)")
sys.exit(0 if ok else 1)
